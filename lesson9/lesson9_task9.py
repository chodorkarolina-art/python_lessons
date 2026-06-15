# Prosty arkusz kalkulacyjny: Używając openpyxl , stwórz plik finanse.xlsx . 
# W pierwszej kolumnie umieść nazwy wydatków (np. "Czynsz", "Jedzenie"), a w drugiej ich wartości. 
# W komórce poniżej wartości oblicz i wstaw sumę wszystkich wydatków, używając formuły Excela (np. =SUM(B1:B2) ).

from openpyxl import Workbook

# tworzymy plik
wb = Workbook()
ws = wb.active

# dane
wydatki = [
    ("Czynsz", 1500),
    ("Jedzenie", 800),
    ("Transport", 300)
]

# wpisywanie danych do arkusza
for i, (nazwa, kwota) in enumerate(wydatki, start=1):
    ws[f"A{i}"] = nazwa
    ws[f"B{i}"] = kwota

# dodanie sumy (Excelowa formuła)
ws[f"A{len(wydatki) + 1}"] = "Suma"
ws[f"B{len(wydatki) + 1}"] = f"=SUM(B1:B{len(wydatki)})"

# zapis pliku
wb.save("finanse.xlsx")